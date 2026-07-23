"""Export a slate — projections + play suggestions — to a formatted workbook.

A thin, deterministic formatter: given the frames the live runner already builds
(game projections, game plays, and optionally prop plays), it writes a styled
``.xlsx`` with a Read Me, one sheet per frame, and a staked total. Values only —
no formulas — so it opens correctly in any viewer with nothing to recalculate.

Formatting is keyed by column name (see the ``*_COLS`` sets), so the runner just
hands over tidy DataFrames with the agreed headers.
"""

from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
NAVY = "1F3864"
_HEADER_FILL = PatternFill("solid", fgColor=NAVY)
_BAND_FILL = PatternFill("solid", fgColor="F2F5FA")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Column formatting, by header name.
_PCT_COLS = {"Model %", "Fair %", "Edge", "Home Win %", "Away Win %", "Stake %", "Over %"}
_MONEY_COLS = {"Stake $"}
_INT_COLS = {"Price"}
_D2_COLS = {"Proj Away", "Proj Home", "Proj Total", "Proj"}
_D1_COLS = {"Fair Total", "Fair Line (Home)", "F5 Total", "Line", "Sample Line"}
_LEFT_COLS = {"Matchup", "Away", "Home", "Player", "Team"}
_TOTAL_COLS = ("Stake %", "Stake $")


def _coerce(value: Any) -> Any:
    """Convert a pandas/numpy cell into a plain type openpyxl accepts."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (int, float)):
        return float(value)
    return str(value)


def _fmt(cell: Any, name: str) -> None:
    cell.font = Font(name=ARIAL, size=10)
    cell.alignment = Alignment(horizontal="left" if name in _LEFT_COLS else "center")
    if name in _PCT_COLS:
        cell.number_format = "0.0%"
    elif name in _MONEY_COLS:
        cell.number_format = "$#,##0.00"
    elif name in _INT_COLS:
        cell.number_format = "+#,##0;-#,##0"
    elif name in _D2_COLS:
        cell.number_format = "0.00"
    elif name in _D1_COLS:
        cell.number_format = "0.0"


def _title(ws: Any, title: str, subtitle: str, ncols: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))
    top = ws.cell(row=1, column=1, value=title)
    top.font = Font(name=ARIAL, bold=True, size=15, color=NAVY)
    ws.row_dimensions[1].height = 21
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ncols, 1))
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.font = Font(name=ARIAL, italic=True, size=10, color="808080")


def _sheet(wb: Workbook, title: str, subtitle: str, df: pd.DataFrame, *, total: bool) -> None:
    ws = wb.create_sheet(title)
    cols = list(df.columns)
    ncols = len(cols)
    _title(ws, title, subtitle, ncols)

    hrow = 4
    for j, name in enumerate(cols, start=1):
        cell = ws.cell(row=hrow, column=j, value=name)
        cell.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for i, (_, rec) in enumerate(df.iterrows()):
        rr = hrow + 1 + i
        for j, name in enumerate(cols, start=1):
            cell = ws.cell(row=rr, column=j, value=_coerce(rec[name]))
            _fmt(cell, name)
            cell.border = _BORDER
            if i % 2 == 1:
                cell.fill = _BAND_FILL

    if total and len(df):
        tr = hrow + 1 + len(df)
        label = ws.cell(row=tr, column=1, value="TOTAL")
        label.font = Font(name=ARIAL, bold=True, color=NAVY)
        for name in _TOTAL_COLS:
            if name not in cols:
                continue
            col = cols.index(name) + 1
            value = float(pd.to_numeric(df[name], errors="coerce").sum())
            cell = ws.cell(row=tr, column=col, value=round(value, 4))
            _fmt(cell, name)
            cell.font = Font(name=ARIAL, bold=True, color=NAVY)
            cell.border = Border(top=Side(style="double", color=NAVY))

    for j, name in enumerate(cols, start=1):
        wide = 34 if name == "Matchup" else 22 if name in ("Player", "Team") else 0
        ws.column_dimensions[get_column_letter(j)].width = wide or max(len(name) + 2, 11)
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)


def _read_me(ws: Any, league: str, generated_at: str, bankroll: float, has_props: bool) -> None:
    _title(ws, "Velocity — Slate", f"{league.upper()} · generated {generated_at}", 2)
    lines = [
        ("Game Projections", "Model fair line per game: projected runs, total, spread, win %."),
        ("Play Suggestions", "Game-market bets clearing the edge threshold, Kelly-staked."),
        ("Prop Suggestions", "Player-prop bets (shown only when a prop board was available)."),
        ("", ""),
        ("Model %", "The model's probability for the bet."),
        ("Fair %", "The de-vigged market probability (both sides)."),
        ("Edge", "Model % - Fair %. Stake % is % of bankroll; Stake $ = Stake % x bankroll."),
        ("Bankroll assumed", None),
        ("", ""),
        ("Note", "A single slate is a sanity check, not proof of edge — the durable signal is"),
        ("", "closing-line value (CLV), validated by backtest. Research tool, not betting advice."),
    ]
    r = 4
    for label, val in lines:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name=ARIAL, bold=bool(label), size=10, color=NAVY)
        if label == "Bankroll assumed":
            bc = ws.cell(row=r, column=2, value=float(bankroll))
            bc.font = Font(name=ARIAL, bold=True, size=11, color=NAVY)
            bc.number_format = "$#,##0"
        elif val:
            vc = ws.cell(row=r, column=2, value=val)
            vc.font = Font(name=ARIAL, size=10, color="404040")
        r += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 86


def _matchups(events: pd.DataFrame) -> dict[str, tuple[str, str]]:
    """``{game_id: (away_name, home_name)}`` from the events frame."""
    return {
        str(g): (str(a), str(h))
        for g, a, h in zip(
            events["game_id"], events["away_team"], events["home_team"], strict=False
        )
    }


def _matchup_label(gid: Any, mm: Mapping[str, tuple[str, str]]) -> str:
    away, home = mm.get(str(gid), ("", ""))
    return f"{away} @ {home}".strip(" @")


_PLAY_COLS = ["Matchup", "Market", "Side", "Line", "Book", "Price",
              "Model %", "Fair %", "Edge", "Stake %", "Stake $"]
_PROP_COLS = ["Matchup", "Player", "Market", "Side", "Line", "Book", "Price",
              "Model %", "Fair %", "Edge", "Stake %", "Stake $"]
_PROJ_COLS = ["Matchup", "Away", "Home", "Proj Away", "Proj Home", "Proj Total",
              "Fair Total", "Fair Line (Home)", "Home Win %", "Away Win %"]


def projections_display(projections: Mapping[str, Any], events: pd.DataFrame) -> pd.DataFrame:
    """Build the Game Projections frame from ``{game_id: GameProjection}`` + events.

    Duck-typed on the projection: it needs ``mu_away``/``mu_home`` and the
    ``fair_total`` / ``fair_spread`` / ``p_home_win`` / ``p_away_win`` helpers.
    """
    mm = _matchups(events)
    rows = []
    for gid, p in projections.items():
        away, home = mm.get(str(gid), ("", ""))
        rows.append({
            "Matchup": _matchup_label(gid, mm), "Away": away, "Home": home,
            "Proj Away": round(float(p.mu_away), 2), "Proj Home": round(float(p.mu_home), 2),
            "Proj Total": round(float(p.mu_away + p.mu_home), 2),
            "Fair Total": round(float(p.fair_total()), 1),
            "Fair Line (Home)": round(float(p.fair_spread()), 1),
            "Home Win %": round(float(p.p_home_win()), 4),
            "Away Win %": round(float(p.p_away_win()), 4),
        })
    return pd.DataFrame(rows, columns=_PROJ_COLS)


def plays_display(frame: pd.DataFrame, events: pd.DataFrame, bankroll: float) -> pd.DataFrame:
    """Map a ``slate_to_frame`` output to the display columns, with team names."""
    mm = _matchups(events)
    if frame.empty:
        return pd.DataFrame(columns=_PLAY_COLS)
    return pd.DataFrame({
        "Matchup": [_matchup_label(g, mm) for g in frame["game_id"]],
        "Market": frame["market"], "Side": frame["side"], "Line": frame["point"],
        "Book": frame["book"], "Price": frame["price"], "Model %": frame["p_model"],
        "Fair %": frame["p_fair"], "Edge": frame["edge"],
        "Stake %": frame["stake"] / bankroll, "Stake $": frame["stake"],
    }, columns=_PLAY_COLS)


def props_display(frame: pd.DataFrame, events: pd.DataFrame, bankroll: float) -> pd.DataFrame:
    """Map a ``prop_slate_to_frame`` output to the display columns, with team names."""
    mm = _matchups(events)
    if frame.empty:
        return pd.DataFrame(columns=_PROP_COLS)
    return pd.DataFrame({
        "Matchup": [_matchup_label(g, mm) for g in frame["game_id"]],
        "Player": frame["player"], "Market": frame["market"], "Side": frame["side"],
        "Line": frame["point"], "Book": frame["book"], "Price": frame["price"],
        "Model %": frame["p_model"], "Fair %": frame["p_fair"], "Edge": frame["edge"],
        "Stake %": frame["stake"] / bankroll, "Stake $": frame["stake"],
    }, columns=_PROP_COLS)


def export_slate_workbook(
    dest: str | Path,
    projections: pd.DataFrame,
    plays: pd.DataFrame,
    props: pd.DataFrame | None = None,
    *,
    league: str,
    generated_at: str,
    bankroll: float = 100.0,
) -> Path:
    """Write the slate workbook to ``dest`` and return its path.

    ``projections`` / ``plays`` / ``props`` are tidy DataFrames with the headers
    the formatter recognizes (see the ``*_COLS`` sets). ``props`` may be ``None``
    or empty, in which case the prop sheet is omitted.
    """
    wb = Workbook()
    _read_me(wb.active, league, generated_at, bankroll, props is not None and not props.empty)
    wb.active.title = "Read Me"
    _sheet(wb, "Game Projections", "Model fair lines", projections, total=False)
    _sheet(wb, "Play Suggestions", "Game-market bets, ranked by edge", plays, total=True)
    if props is not None and not props.empty:
        _sheet(wb, "Prop Suggestions", "Player-prop bets, ranked by edge", props, total=True)
    dest = Path(dest)
    wb.save(str(dest))
    return dest
